import ttkbootstrap as tb
import tkinter as tk
import openpyxl
import pandas as pd
import os, sys

from notebookconfigurations.createnotebook import clear_all_notebooks
# Declare file path to the lookup table.
assaysPath = r"\\mfad.mfroot.org\rchapp\Tox\Dept\GC\DOA Repeat Sheets\RTS ONLY - Repeat Sheets and Materials\Repeat Sheet Instrument and Assay List.xlsx"


# Function to retrieve the assay and instrument data
def get_Assay_and_Instrument():
    assayLookUpTable = openpyxl.load_workbook(assaysPath)
    sheet = assayLookUpTable['Assay']
    test_codes = []
    instruments = {}

    # Assuming the "Test Code" is in the first column and instruments are in columns 2, 3, and 4
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        test_code = row[0]  # "Test Code"
        instruments_list = row[1:4]  # Columns 2, 3, and 4 for instruments
        
        # Add the test codes for the Assay dropdown
        if test_code:
            test_codes.append(test_code.upper())
        
        # Create a dictionary where the key is the Test Code and value is a list of instruments
        if test_code not in instruments:
            instruments[test_code] = []
        
        # Add the instruments to the list, ensuring no duplicates
        for instrument in instruments_list:
            if instrument:  # Skip blank values
                if instrument not in instruments[test_code]:
                    instruments[test_code].append(instrument.upper())

    # Create a dictionary to pair each unique instrument with its corresponding test codes
    instrument_assay_mapping = {}
    for test_code, instrument_list in instruments.items():
        for instrument in instrument_list:
            if instrument not in instrument_assay_mapping:
                instrument_assay_mapping[instrument] = []
            instrument_assay_mapping[instrument].append(test_code)

    # Flatten the instrument list so each instrument is a separate option in the Combobox
    all_instruments = list(instrument_assay_mapping.keys())
    # Sort both lists alphabetically
    test_codes = sorted(test_codes)
    all_instruments = sorted(all_instruments)
    return test_codes, all_instruments, instrument_assay_mapping

# Create the function designed to return all of the necessary/key components to dynamically assign the requirements for creating the PM form.
def PM_QUERY(assay,instrument):
    # add the instrument name into the path too.
    assay = assay.upper()
    instrument = instrument.upper()
    # If the program is running as a script
    #project_directory = fr"{os.getcwd()}\QUERY_FILES"
    #print(f'Main directory: {project_directory}')
    #pm_query_path = fr"{project_directory}\{assay}_{instrument}_PM_FORM_QUERY.xlsx"

    pm_query_path = fr"\\mfad.mfroot.org\rchapp\Tox\Dept\ITTS Files\Projects\PM Project\PM Application\QUERY_FILES\{assay}_{instrument}_PM_FORM_QUERY.xlsx"
    #print(f"File for assay found {pm_query_path}")
    # Check if the file exists
    if not os.path.exists(pm_query_path):
        print(f"File for assay '{assay}' not found, defaulting to the base file path.")
        pm_query_path = r"\\mfad.mfroot.org\rchapp\Tox\Dept\ITTS Files\Projects\PM Project\PM Application\QUERY_FILES\PM_FORM_QUERY.xlsx"
    #else:
        #pm_query_path = r"\\mfad.mfroot.org\rchapp\Tox\Dept\ITTS Files\Projects\PM Project\PM Application\QUERY_FILES\PM_FORM_QUERY.xlsx"

    # load the workbook.
    my_pm_table = openpyxl.load_workbook(pm_query_path,read_only=True)
    # access the first page for creating the daily PM form
    sheets = my_pm_table['Daily']
    
    # declare my lists to track all necessary parameters.
    task_type = []
    ui_type = []
    task_name = []
    operator_symbol = []
    cutoffValue = []
    margin_value = []
    unit_value = []
    optional_value = []

    # loop through the rows to retrieve data.
    for row in sheets.iter_rows(min_row=2,min_col=1,max_col=8):
        task_type.append(row[0].value)          # Column A: Task Type
        ui_type.append(row[1].value)            # Column B: UI Type
        task_name.append(row[2].value)          # Column C: Task Name
        operator_symbol.append(row[3].value)    # Column D: Operator
        cutoffValue.append(row[4].value)       # Column E: Cutoff Value
        margin_value.append(row[5].value)       # Column F: Margin Value
        unit_value.append(row[6].value)         # Column G: units
        optional_value.append(row[7].value)     # Column H: optional status
    my_pm_table.close()
    return task_type, ui_type, task_name, operator_symbol, cutoffValue, margin_value, unit_value, optional_value

# Function for retrieving the approved assays based upon the instrument selection
# update_functions.py

def update_instruments(event, root, instrument_var, assay_drop_down_combo, instrument_assay_mapping):
    selected_instrument = instrument_var.get()
    clear_all_notebooks(root)
    # Update the assay dropdown values based on the selected instrument
    if selected_instrument in instrument_assay_mapping:
        # Get the corresponding assays for the selected instrument
        assays_for_instrument = instrument_assay_mapping[selected_instrument]
        assay_drop_down_combo['values'] = assays_for_instrument
        assay_drop_down_combo.set("")  # Clear any existing selection
    else:
        assay_drop_down_combo['values'] = []  # No assays available for selected instrument
        assay_drop_down_combo.set("")  # Clear selection

